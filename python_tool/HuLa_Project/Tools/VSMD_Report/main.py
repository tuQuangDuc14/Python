# lib parser
import argparse
# lib logging
import logging
# sub process
import subprocess
# get time
import datetime
logger = logging.getLogger(__name__)
logging.basicConfig(
    level=logging.DEBUG,
    handlers=[logging.StreamHandler()],
    datefmt='%Y-%m-%d %H:%M:%S',
    format='[%(asctime)s %(levelname)s] --> %(module)s: %(message)s'
)

# provides functions for creating and removing a directory (folder), fetching its contents, changing and identifying the current directory
import os
 # remove the directory
import shutil
# get xml
import xml.etree.ElementTree as ET
# import
from Common.common import Make_Dir, CommonStep, StepStatus
# phân tích cú pháp HTML và XML
# pip install beautifulsoup4
from bs4 import BeautifulSoup

# sudo pip3 install openpyxl
# excell
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import worksheet
from openpyxl.worksheet import table
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

from openpyxl.styles.borders import Border, Side


def Read_argument():
    # init parser
    parser = argparse.ArgumentParser(description="write args")
    # Options for tool
    parser.add_argument("-m", "--module", help="Name module", nargs='+', default=None)
    return parser.parse_args()

# chứa các biến đầu vào input
class mConfig():
    #Tool_Dir
    #get current path of folder
    Tool_Dir = os.getcwd()
    # Normalize the specified path
    # using os.path.normpath() method
    Output_Report = os.path.normpath(f'{Tool_Dir}/Output')
    #get from xml
    # Get path VSMD_Report
    Tree = ET.parse('UserConfig.xml')
    Root = Tree.getroot()
    # Get compiler
    EB_Tresos_Dir = Root[0][0].text
    Product = Root[0][1].text
    Derivatives = Root[0][2].text
    Is_Module_bsw00347 = Root[0][3].text

class mRunTool(CommonStep):
    def __init__(self, module):
        super().__init__(module)
        self.Module = module
        self.mConfig = mConfig()
        self.List_Derivatives = []
        for Derivatives in self.mConfig.Derivatives.split(","):
            self.List_Derivatives.append(Derivatives)
        self.Tresos_Dir = self.mConfig.EB_Tresos_Dir
        self.Bat_File_Tresos =  os.path.normpath(self.Tresos_Dir  + '/bin/tresos_cmd.bat')

        if self.mConfig.Is_Module_bsw00347 == "True":
            self.EPD_File_AUTOSAR = self.Tresos_Dir + "/autosar/4.4.0/" + self.Module.split('_')[0] + '.epd'
        else:
            # self.EPD_File_AUTOSAR = os.path.normpath(os.path.join( self.Tresos_Dir, "autosar/4.4.0",f'{self.Module}.epd'))
            self.EPD_File_AUTOSAR = os.path.normpath(self.Tresos_Dir + "/autosar/4.4.0/" + f'{self.Module}.epd')

        self.My_EPD_File_1 = self.Tresos_Dir + "/plugins/" + self.Module + "_HuLa/autosar/" + self.Module + ".epd" 
        self.My_EPD_File_2 = self.Tresos_Dir + "/plugins/" + self.Module + "_HuLa/autosar/" + self.Module + "_" + self.mConfig.Derivatives + ".epd" 
        self.XDM_File = self.Tresos_Dir + "/plugins/" + self.Module + "_HuLa/config/" + self.Module + ".xdm" 
        self.Create_Excell()

    # init excel
    def Create_Excell(self):
        # open 1 cái excel và gán cho self.Template_Report_WorkBook
        self.Template_Report_WorkBook = openpyxl.load_workbook('Template/VSMD_Report_Template.xlsx')
        # truy cap vao sheet "Title" và gán cho self.Title_Sheet
        self.Title_Sheet = self.Template_Report_WorkBook["Title"]
        # truy cập vào ô I7 trên sheet "Title" và gán giá trị cho ô đó
        self.Title_Sheet['I7'] = self.Module.upper() 
        # the current date
        Current_Time = datetime.datetime.now()
        # đưa về định dạng nam tháng ngay
        date = Current_Time.strftime("%Y-%m-%d")
        self.Title_Sheet['I8'] = date
        # truy cap vao sheet "VSMD" và gán cho self.VSMD_Sheet
        self.VSMD_Sheet = self.Template_Report_WorkBook["VSMD"]
        self.Summary_Sheet = self.Template_Report_WorkBook["Summary"]
        # Border (đường viền) cho các ô minh muốn
        self.MyBoder = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        # self.Title_Sheet['I10'].border = self.MyBoder
        # self.Template_Report_WorkBook.save('Template/VSMD_Report_Template.xlsx') 
        # Create a dictionary
        self.Violations_Count_Dict = dict()
         
    def Collect_Violation_And_Export_Excel(self):
        self.My_Row = 2
        for Derivative in self.List_Derivatives:
            self.Violations_Count_Dict[Derivative] = dict()
            # get file .epd from plugins
            if os.path.exists(self.My_EPD_File_1):
                self.EPD_File_Name = self.Module + ".epd"
            if os.path.exists(self.My_EPD_File_2):
                self.EPD_File_Name = self.Module + "_" + Derivative + ".epd"
            else:
                self.EPD_File_Name = self.Module + "_" + Derivative + ".epd"
            # Get path of epd file
            self.EPD_File = self.Tresos_Dir + "/plugins/" + self.Module + "_HuLa/autosar/" + self.EPD_File_Name
            # print(self.EPD_File)
            # contain data when vsmd check done
            self.File_XDM_Check_Output = f'Output/{self.Module}/xdm_check_' + Derivative.lower() + '.xml'
            self.File_EPD_Check_Output = f'Output/{self.Module}/epd_check_' + Derivative.lower() + '.xml'
            # print(self.File_XDM_Check_Output)
            # print(self.File_EPD_Check_Output)
            # CMD = tresos_cmd.bat + -Dtarget=CortexM -Dderivate=HULA legacy vsmdcheck + file epd of autosar(autosar\4.4.0\Gpio.epd) + file xdm of module(plugins /config/Gpio.xdm) + '@xdm asc:4.4.0 -xml '+ File_XDM_Check_Output
            self.CMD_Check_XDM = self.Bat_File_Tresos + ' ' + '-Dtarget=CortexM -Dderivate=' + self.mConfig.Product.upper() + ' -Dsubderivative=' + Derivative.lower() + ' legacy vsmdcheck ' + self.EPD_File_AUTOSAR + ' ' + self.XDM_File +'@xdm asc:4.4.0 -xml ' + self.File_XDM_Check_Output
            # print(self.CMD_Check_XDM)
            self.CMD_Check_EPD = self.Bat_File_Tresos + ' ' + '-Dtarget=CortexM -Dderivate=' + self.mConfig.Product.upper() + ' -Dsubderivative=' + Derivative.lower() + ' legacy vsmdcheck ' + self.EPD_File_AUTOSAR + ' ' + self.EPD_File +' asc:4.4.0 -xml ' + self.File_EPD_Check_Output
            # run cmd check vsmd 
            logging.info(f'START CHECKING XDM FILE {self.XDM_File} ...................')
            # p = subprocess.Popen(["python", "--help"], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            # output, errors = p.communicate()
            # print(output)
            Output_XDM_Check = subprocess.Popen(self.CMD_Check_XDM)
            Output_XDM_Check.communicate()
            logging.info(f'START CHECKING EPD FILE {self.EPD_File} ...................')
            Output_EPD_Check = subprocess.Popen(self.CMD_Check_EPD)
            Output_EPD_Check.communicate()
            # ret, log = CommonStep.RunCmd(self.CMD_Check_XDM)
            # print(log)
            #find tất cả các tag có tên là "violation"
            XDM_Violations = BeautifulSoup(open(self.File_XDM_Check_Output, encoding="utf8"), features="xml").find_all('violation')
            # bo qua các req nay
            Ignore_List = ["EcucSws_1014", "EcucSws_2038_2040_ASR41"]
            self.Violations_Count_Dict[Derivative][self.Module + '.xdm'] = 0
            for Violation in XDM_Violations:
                RuleId = Violation.find('ruleid').text
                if not RuleId in Ignore_List:
                    self.Violations_Count_Dict[Derivative][self.Module + '.xdm'] += 1
                    self.VSMD_Sheet['A' + str(self.My_Row)].border = self.MyBoder
                    self.VSMD_Sheet['B' + str(self.My_Row)].border = self.MyBoder
                    self.VSMD_Sheet['C' + str(self.My_Row)].border = self.MyBoder
                    self.VSMD_Sheet['D' + str(self.My_Row)].border = self.MyBoder
                    self.VSMD_Sheet['E' + str(self.My_Row)].border = self.MyBoder
                    self.VSMD_Sheet['F' + str(self.My_Row)].border = self.MyBoder
                    self.VSMD_Sheet['G' + str(self.My_Row)].border = self.MyBoder
                    self.VSMD_Sheet['H' + str(self.My_Row)].border = self.MyBoder
                    self.VSMD_Sheet['A' + str(self.My_Row)] = self.Module + '.xdm'
                    self.VSMD_Sheet['B' + str(self.My_Row)] = Derivative.lower()
                    self.VSMD_Sheet['C' + str(self.My_Row)] = Violation.find('ruleid').text
                    self.VSMD_Sheet['D' + str(self.My_Row)] = Violation.find('description').text
                    self.VSMD_Sheet['E' + str(self.My_Row)] = Violation.find('message').text
                    if(None != Violation.find('vsmdNode')):
                        self.VSMD_Sheet['F' + str(self.My_Row)] = Violation.find('vsmdNode').text
                    if(None != Violation.find('stmdNode')):
                        self.VSMD_Sheet['G' + str(self.My_Row)] = Violation.find('stmdNode').text
                    self.My_Row += 1
            
            # find tất cả các các tag có tên violations
            EPD_Violations = BeautifulSoup(open(self.File_EPD_Check_Output, encoding="utf8"), features="xml").find_all('violation')

            self.Violations_Count_Dict[Derivative][self.EPD_File_Name] = 0
            for Violation in EPD_Violations:
                RuleId = Violation.find('ruleid').text
                if not RuleId in Ignore_List:
                    self.Violations_Count_Dict[Derivative][self.EPD_File_Name] += 1
                    self.VSMD_Sheet['A' + str(self.My_Row)].border = self.MyBoder
                    self.VSMD_Sheet['B' + str(self.My_Row)].border = self.MyBoder
                    self.VSMD_Sheet['C' + str(self.My_Row)].border = self.MyBoder
                    self.VSMD_Sheet['D' + str(self.My_Row)].border = self.MyBoder
                    self.VSMD_Sheet['E' + str(self.My_Row)].border = self.MyBoder
                    self.VSMD_Sheet['F' + str(self.My_Row)].border = self.MyBoder
                    self.VSMD_Sheet['G' + str(self.My_Row)].border = self.MyBoder
                    self.VSMD_Sheet['H' + str(self.My_Row)].border = self.MyBoder
                    self.VSMD_Sheet['A' + str(self.My_Row)] = self.EPD_File_Name
                    self.VSMD_Sheet['B' + str(self.My_Row)] = Derivative.lower()
                    self.VSMD_Sheet['C' + str(self.My_Row)] = Violation.find('ruleid').text
                    self.VSMD_Sheet['D' + str(self.My_Row)] = Violation.find('description').text
                    self.VSMD_Sheet['E' + str(self.My_Row)] = Violation.find('message').text
                    if(None != Violation.find('vsmdNode')):
                        self.VSMD_Sheet['F' + str(self.My_Row)] = Violation.find('vsmdNode').text
                    if(None != Violation.find('stmdNode')):
                        self.VSMD_Sheet['G' + str(self.My_Row)] = Violation.find('stmdNode').text
                    self.My_Row += 1
            os.remove(self.File_EPD_Check_Output)
            os.remove(self.File_XDM_Check_Output)

        self.My_Row = 2
        self.Count = 0
        for Derivative in self.List_Derivatives:
            for File in self.Violations_Count_Dict[Derivative].keys():
                self.Summary_Sheet['A' + str(self.My_Row)].border = self.MyBoder
                self.Summary_Sheet['B' + str(self.My_Row)].border = self.MyBoder
                self.Summary_Sheet['C' + str(self.My_Row)].border = self.MyBoder
                self.Summary_Sheet['A' + str(self.My_Row)] = File
                self.Summary_Sheet['B' + str(self.My_Row)] = Derivative
                self.Summary_Sheet['C' + str(self.My_Row)] = self.Violations_Count_Dict[Derivative][File]
                self.Count += self.Violations_Count_Dict[Derivative][File]
                self.My_Row += 1
            self.Summary_Sheet['A' + str(self.My_Row)].border = self.MyBoder
            self.Summary_Sheet['B' + str(self.My_Row)].border = self.MyBoder
            self.Summary_Sheet['C' + str(self.My_Row)].border = self.MyBoder
            self.Summary_Sheet['B' + str(self.My_Row)] = 'Totals:'
            self.Summary_Sheet['C' + str(self.My_Row)] = self.Count
        
        Output_Report_Temp = f'{self.mConfig.Output_Report}/{self.Module}'
        # save new excel vao folder output
        self.Template_Report_WorkBook.save(Output_Report_Temp + '/VSMD_Report_' + self.Module.upper() + '_ASR_440.xlsx')
    
    def Start(self):
        # Remove and create folder Output
        Output_Report_Temp = f'{self.mConfig.Output_Report}/{self.Module}'
        # print(Output_Report_Temp)
        # Check whether the specified
        # path exists or not
        isExist = os.path.exists(Output_Report_Temp)
        if isExist:
            # remove the directory
            shutil.rmtree(Output_Report_Temp)
        # create the directory
        Make_Dir(Output_Report_Temp)

        self.Result = StepStatus.FAILED
        self.LetStart()
        #to do
        self.Collect_Violation_And_Export_Excel()

        self.Result = StepStatus.SUCCEEDED
        self.LetEnd()

if __name__ == '__main__':
    Argument = Read_argument()
    logging.info("  Argument parameter is %s ",Argument.module[0])
    RunNew = mRunTool(Argument.module[0])
    RunNew.Start()



